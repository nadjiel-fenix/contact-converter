import openpyxl
import re
import argparse
from pathlib import Path


def extract_phone(text: str) -> str:
    """
    Valida se uma string representa um número de telefone, e retorna
    sua versão padronizada no formato E.164 (sem o +).
    Se a string não é um número de telefone, entretanto, esta
    função retorna uma string vazia.
    """

    # Remove caracteres não numéricos.
    result = re.sub(r"\D", "", text)

    # Remove 0 inicial, se presente.
    if result.startswith("0"):
        result = result[1:]

    # Remove código do Brasil, se houver.
    if result.startswith("55") and len(result) > 11:
        result = result[2:]

    # Valida tamanho do número: 10 (fixo) or 11 (celular).
    if len(result) in [10, 11]:
        return "55" + result

    return result if len(result) in [10, 11] else ""


def convert(input_path: str, output_path: str):
    """
    Converte os dados do arquivo de contatos do DropDesk apontado
    pelo input_path (.xlsx) para um novo arquivo no caminho output_path
    com o formato dos dados esperados pelo AtendeChat (também .xlsx).
    """

    # Carregar workbook do DropDesk.
    in_wb = openpyxl.load_workbook(input_path)
    # Carrega o worksheet padrão, que se
    # espera que contenha os dados exportados.
    in_ws = in_wb.active

    # Extrai os headers
    headers = [
        cell.value.strip() if cell.value else ""
        for cell in next(in_ws.iter_rows(min_row=1, max_row=1))
    ]
    header_map = {h: i for i, h in enumerate(headers)}

    # Prepara novo workbook para output.
    out_wb = openpyxl.Workbook()
    out_ws = out_wb.active
    # Adiciona headers esperados pelo AtendeChat
    out_ws.append(["Nome", "Telefone"])

    for row in in_ws.iter_rows(min_row=2):
        nome = row[header_map["Nome"]].value or ""
        empresa = row[header_map["Empresa"]].value or ""
        telefone = row[header_map["Telefone"]].value or ""

        nome_final = nome.strip() if nome.strip() else empresa.strip()
        telefone_final = extract_phone(telefone) or extract_phone(empresa)

        out_ws.append([nome_final, telefone_final])

    out_wb.save(output_path)


def main() -> None:
    """
    Interpreta os argumentos e executa a conversão.
    """

    parser = argparse.ArgumentParser(
        description="Converte contatos do DropDesk para o formato do AtendeChat."
    )
    parser.add_argument(
        "input",
        help="Caminho do arquivo de input com os contatos do DropDesk (.xlsx)",
        type=Path,
    )
    parser.add_argument(
        "output",
        help="Caminho do arquivo onde salvar os contatos no formato do AtendeChat (.xlsx)",
        type=Path,
    )
    args = parser.parse_args()

    convert(args.input, args.output)


if __name__ == "__main__":
    main()
